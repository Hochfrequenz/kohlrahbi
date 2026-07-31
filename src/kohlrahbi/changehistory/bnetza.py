"""Module to download and handle BNetzA change-history documents (PDF, Office and HTML)."""

import asyncio
import io
import logging
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from urllib.parse import unquote, urldefrag, urljoin, urlparse

import docx
import httpx
import openpyxl  # type: ignore[import-untyped]
import pandas as pd
import pdfplumber
from bs4 import BeautifulSoup
from bs4.element import Tag
from openpyxl.styles import Alignment  # type: ignore[import-untyped]

from kohlrahbi.changehistory import get_change_history_table

logger = logging.getLogger(__name__)

# BNetzA occasionally blocks clients without a browser-like User-Agent; set one defensively.
_HTTP_HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; kohlrahbi change-history scraper)"}

# Map a detected document kind to the file extension we store it under.
_EXTENSION_BY_KIND = {"pdf": ".pdf", "xlsx": ".xlsx", "docx": ".docx", "html": ".html"}


def clean_filename(text: str) -> str:
    """
    Clean up the filename by removing file size information and special characters.

    Args:
        text: The text to clean up

    Returns:
        A clean filename
    """
    # Remove the file size pattern (pdf / X MB) or (pdf / X KB)
    text = re.sub(r"\s*\(pdf\s*\/\s*\d+(?:\.\d+)?\s*[KMG]B\)", "", text)

    # Replace problematic characters
    text = text.replace("/", "_").replace(" ", "_")

    # Add .pdf extension if not present
    if not text.lower().endswith(".pdf"):
        text += ".pdf"

    return text


def _resolve_base_url(page_url: str, soup: BeautifulSoup) -> str:
    """
    Determine the base URL used to resolve relative links on the page.

    BNetzA pages set ``<base href="/">`` and emit document links that are relative *without*
    a leading slash (e.g. ``DE/Beschlusskammern/...``), so they must be resolved against the
    site root rather than the document's directory.
    """
    base_tag = soup.find("base")
    if isinstance(base_tag, Tag):
        base_href = base_tag.get("href")
        if isinstance(base_href, str) and base_href:
            return urljoin(page_url, base_href)
    return page_url


def _is_download_link(anchor: Tag) -> bool:
    """
    Decide whether an anchor points to a downloadable BNetzA document.

    Download links are marked either by a CSS class (``downloadLink``/``Publication``) or by a
    tell-tale query/fragment (``__blob=publicationFile`` or ``#download=1``). Matching on any of
    these catches PDFs, Office documents and the ``.html``-named documents alike.
    """
    href = anchor.get("href")
    if not isinstance(href, str) or not href:
        return False
    class_attr = anchor.get("class")
    class_str = " ".join(class_attr) if isinstance(class_attr, list) else (class_attr or "")
    if "downloadLink" in class_str or "Publication" in class_str:
        return True
    return "__blob=publicationFile" in href or "#download=1" in href


def extract_document_links(page_html: str, page_url: str) -> list[tuple[str, str]]:
    """
    Parse all downloadable document links out of a BNetzA page's HTML.

    Returns a list of ``(absolute_url, link_text)`` tuples, de-duplicated by absolute URL
    (ignoring the fragment). Relative hrefs are resolved against the page's ``<base>`` element.
    """
    soup = BeautifulSoup(page_html, "html.parser")
    base_url = _resolve_base_url(page_url, soup)

    links: list[tuple[str, str]] = []
    seen: set[str] = set()
    for anchor in soup.find_all("a"):
        if not isinstance(anchor, Tag) or not _is_download_link(anchor):
            continue
        href = anchor.get("href")
        assert isinstance(href, str)  # guaranteed by _is_download_link
        absolute_url = urldefrag(urljoin(base_url, href)).url
        if absolute_url in seen:
            continue
        seen.add(absolute_url)
        links.append((absolute_url, anchor.get_text(strip=True)))
    return links


async def get_document_links(url: str) -> list[tuple[str, str]]:
    """
    Fetch all downloadable document links from the given BNetzA website URL.

    Returns a list of ``(absolute_url, link_text)`` tuples, de-duplicated by absolute URL
    (ignoring the fragment). Unlike the previous implementation this is not limited to ``.pdf``
    links: newer BNetzA pages serve most EDIFACT documents as ``.html``-named downloads and as
    Office files, all of which are included here.

    Args:
        url: The BNetzA page URL to scrape for document links.
    """
    async with httpx.AsyncClient(timeout=httpx.Timeout(30.0), follow_redirects=True, headers=_HTTP_HEADERS) as client:
        response = await client.get(url)
        response.raise_for_status()
        page_html = response.text

    links = extract_document_links(page_html, url)
    logger.info("Found %d document links to download", len(links))
    return links


def _filename_stem_from_url(url: str) -> str:
    """
    Derive a stable filename stem from the document URL (not the anchor text).

    Using the URL avoids collisions between documents that share a human-readable label and
    yields the real document name (e.g. ``UTILMD_AHB_Strom_2_3_Konsultationsfassung_20260731``).
    The extension is intentionally dropped here; the real one is chosen after download based on
    the actual content type.
    """
    path = urlparse(url).path
    name = unquote(path.rsplit("/", 1)[-1])
    stem = name.rsplit(".", 1)[0] if "." in name else name
    return stem or "document"


def _kind_from_content_type(content_type: str | None) -> str:
    """Map an HTTP ``Content-Type`` to a document kind (fallback when magic bytes are unclear)."""
    if not content_type:
        return "unknown"
    normalized = content_type.split(";")[0].strip().lower()
    if "spreadsheetml" in normalized:
        return "xlsx"
    if "wordprocessingml" in normalized:
        return "docx"
    simple = {
        "application/pdf": "pdf",
        "text/html": "html",
        "application/xhtml+xml": "html",
    }
    return simple.get(normalized, "unknown")


def detect_document_type(content: bytes, content_type: str | None = None) -> str:
    """
    Classify a downloaded document by its content.

    BNetzA serves many documents under a ``.html`` filename even though the body is a PDF, so the
    real type must be detected from the bytes (with the HTTP ``Content-Type`` as a fallback)
    rather than trusted from the URL extension.

    Returns one of ``"pdf"``, ``"xlsx"``, ``"docx"``, ``"html"`` or ``"unknown"``.
    """
    if content[:4] == b"%PDF":
        return "pdf"
    if content[:2] == b"PK":
        try:
            names = zipfile.ZipFile(io.BytesIO(content)).namelist()
        except zipfile.BadZipFile:
            names = []
        if any(name.startswith("xl/") for name in names):
            return "xlsx"
        if any(name.startswith("word/") for name in names):
            return "docx"
    stripped = content[:64].lstrip().lower()
    if stripped.startswith(b"<!doctype html") or stripped.startswith(b"<html"):
        return "html"
    return _kind_from_content_type(content_type)


@dataclass
class DownloadResult:
    """Outcome of downloading a single BNetzA document."""

    url: str
    text: str
    stem: str
    kind: str  # pdf / xlsx / docx / html / unknown
    path: Path | None
    status: str  # "downloaded", "skipped", "failed"
    error: str | None = None


def _find_existing_download(target_dir: Path, stem: str) -> Path | None:
    """Return an already-downloaded file for this stem, if any (any extension)."""
    for candidate in sorted(target_dir.glob(f"{stem}.*")):
        if candidate.is_file():
            return candidate
    return None


async def download_document(
    client: httpx.AsyncClient, url: str, text: str, target_dir: Path, semaphore: asyncio.Semaphore
) -> DownloadResult:
    """
    Download a single document, detect its real type, and store it with the correct extension.

    Skips the download when a file for the same stem already exists on disk (any extension).
    """
    stem = _filename_stem_from_url(url)

    existing = _find_existing_download(target_dir, stem)
    if existing is not None:
        logger.info("File %s already exists, skipping download...", existing.name)
        kind = next((k for k, ext in _EXTENSION_BY_KIND.items() if ext == existing.suffix.lower()), "unknown")
        return DownloadResult(url=url, text=text, stem=stem, kind=kind, path=existing, status="skipped")

    try:
        async with semaphore:
            logger.info("Downloading %s from %s", stem, url)
            response = await client.get(url)
            response.raise_for_status()
            content = response.content
    except httpx.HTTPError as error:
        logger.error("Failed to download %s from %s: %s", stem, url, str(error))
        return DownloadResult(
            url=url, text=text, stem=stem, kind="unknown", path=None, status="failed", error=str(error)
        )

    kind = detect_document_type(content, response.headers.get("content-type"))
    extension = _EXTENSION_BY_KIND.get(kind, Path(urlparse(url).path).suffix or ".bin")
    target_path = target_dir / f"{stem}{extension}"
    try:
        target_path.write_bytes(content)
    except OSError as error:
        logger.error("Failed to write %s: %s", target_path.name, str(error))
        return DownloadResult(url=url, text=text, stem=stem, kind=kind, path=None, status="failed", error=str(error))

    logger.info("Successfully downloaded %s (%s, %d bytes)", target_path.name, kind, len(content))
    return DownloadResult(url=url, text=text, stem=stem, kind=kind, path=target_path, status="downloaded")


def find_change_history_page(pdf: pdfplumber.pdf.PDF) -> int:
    """
    Find the page number where Änderungshistorie starts by checking the table of contents.

    Args:
        pdf: The opened PDF document

    Returns:
        The 0-based page index where Änderungshistorie starts, or -1 if not found
    """
    # Check first few pages for table of contents
    for page_idx, page in enumerate(pdf.pages[:4]):  # Usually TOC is in first few pages
        text = page.extract_text() or ""
        # Look for "Änderungshistorie" followed by a page number
        matches = re.finditer(r"Änderungshistorie[.\s]+(\d+)", text)
        for match in matches:
            # Convert 1-based page number to 0-based index
            page_num = int(match.group(1)) - 1
            logger.debug("Found Änderungshistorie in TOC on page %d, pointing to page %d", page_idx + 1, page_num + 1)
            return page_num

    # Fallback: search through all pages
    for i, page in enumerate(pdf.pages):
        if "Änderungshistorie" in (page.extract_text() or ""):
            logger.debug("Found Änderungshistorie text on page %d", i + 1)
            return i

    logger.debug("No Änderungshistorie section found in any page")
    return -1


def _merge_columns(cells: list[str | None]) -> str:
    """Merge multiple cells into one, joining non-empty values with newlines."""
    parts = [str(c) for c in cells if c is not None and str(c).strip()]
    return "\n".join(parts) if parts else ""


def normalize_table_columns(table: list[list[str | None]]) -> list[list[str | None]]:
    """
    Normalize tables with 10 columns (e.g. Allgemeine Festlegungen) to the standard 6-column layout.

    The 10-column layout from pdfplumber looks like:
        [Änd-ID, Ort, _, Fehlerkorrektur/Änderung, _, _, _, _, Grund der Anpassung, Status]
        [_, _, _, Bisher, _, _, Neu, _, _, _]

    This maps to 6 logical columns:
        [Änd-ID, Ort, Bisher, Neu, Grund der Anpassung, Status]
    """
    if not table or len(table[0]) != 10:
        return table

    normalized = []
    for row in table:
        normalized.append(
            [
                row[0],  # Änd-ID
                row[1],  # Ort
                _merge_columns(row[2:5]),  # Bisher (cols 2-4)
                _merge_columns(row[5:8]),  # Neu (cols 5-7)
                row[8],  # Grund der Anpassung
                row[9],  # Status
            ]
        )
    return normalized


def _merge_row_into(target: list[str], source: list[str]) -> None:
    """Merge non-empty values from source into target, joining with newlines."""
    for i, value in enumerate(source):
        if value:
            if target[i]:
                target[i] = f"{target[i]}\n{value}"
            else:
                target[i] = value


def clean_table_data(table: list[list[str | None]]) -> list[list[str]]:
    """
    Clean up the table data by merging related rows before converting to DataFrame.
    Rows with empty first column should be merged with the row above.

    Args:
        table: Raw table data from PDF

    Returns:
        Cleaned table data with merged rows
    """

    if len(table) < 2:
        logger.warning("Table has insufficient rows (%d), need at least 2", len(table))
        return []

    # Get headers and sub-headers, converting None to empty strings
    headers = [str(cell) if cell is not None else "" for cell in table[0]]
    sub_header = [str(cell) if cell is not None else "" for cell in table[1]]

    # Start with headers and sub-headers
    result: list[list[str]] = [headers]
    result.append(sub_header)

    current_row: list[str] | None = None

    # Process each data row (skip header and sub-header)
    for raw_row in table[2:]:
        # Convert None to empty strings in current row
        row: list[str] = [str(cell) if cell is not None else "" for cell in raw_row]
        # If first column (Änd-ID) is empty, merge with previous row
        if not row[0] and current_row is not None:
            _merge_row_into(current_row, row)
        elif row[0]:
            # If we have a previous row, add it to results
            if current_row is not None:
                result.append(current_row)
            # Start new row
            current_row = row.copy()

    # Add the last row if exists
    if current_row is not None:
        result.append(current_row)

    logger.debug("Cleaned table data: %d rows (including headers)", len(result))
    return result


def _is_change_history_header(first_cell: str | None) -> bool:
    """
    Decide whether a table's first cell marks the start of an Änderungshistorie table.

    The header cell is sometimes split across lines (``"Änd-\\nID"``) or annotated
    (``"Änd-ID\\n…"``), so newlines are stripped and a prefix match is used. This is a strict
    superset of an exact ``"Änd-ID"`` match.
    """
    return (first_cell or "").replace("\n", "").strip().startswith("Änd-ID")


def extract_change_history(pdf_path: Path) -> pd.DataFrame:
    """
    Extract the Änderungshistorie table from a PDF file.
    Specifically looks for tables that start with 'Änd-ID'.
    First finds the page number from table of contents.

    Args:
        pdf_path: Path to the PDF file

    Returns:
        DataFrame containing the change history table
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            logger.debug("Successfully opened PDF %s with %d pages", pdf_path.name, len(pdf.pages))
            # First find the page containing Änderungshistorie from table of contents
            change_history_page = find_change_history_page(pdf)

            if change_history_page == -1:
                logger.warning("No Änderungshistorie section found in %s", pdf_path.name)
                return pd.DataFrame()

            logger.info(
                "Found Änderungshistorie section in %s starting at page %d", pdf_path.name, change_history_page + 1
            )

            all_rows: list[list[str | None]] = []

            # Scan pages from the Änderungshistorie page onwards, collecting raw rows
            for page in pdf.pages[change_history_page:]:
                tables = page.extract_tables()
                logger.debug("Page %d has %d tables", page.page_number + 1, len(tables))

                for table_idx, table in enumerate(tables):
                    if not table or not table[0]:
                        continue

                    if _is_change_history_header(table[0][0]):
                        logger.info(
                            "Found change history table in %s on page %d (table %d)",
                            pdf_path.name,
                            page.page_number + 1,
                            table_idx + 1,
                        )
                        # Normalize 10-column tables to 6 columns
                        normalized_table = normalize_table_columns(table)
                        if not all_rows:
                            # First table: keep header and sub-header
                            all_rows.extend(normalized_table)
                        else:
                            # Subsequent tables: skip header (row 0) and sub-header (row 1)
                            all_rows.extend(normalized_table[2:])

            if all_rows:
                cleaned_table = clean_table_data(all_rows)
                if cleaned_table:
                    df = pd.DataFrame(cleaned_table[1:], columns=cleaned_table[0])
                    logger.info("Extracted %d rows of change history data from %s", len(df), pdf_path.name)
                    return df

            logger.warning("No change history table found in %s", pdf_path.name)
            return pd.DataFrame()
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("Failed to open PDF %s: %s", pdf_path.name, str(e))
        return pd.DataFrame()


def _cell_text(cell: object) -> str:
    """Normalize a raw cell value (openpyxl/BeautifulSoup) to a stripped string."""
    return "" if cell is None else str(cell).strip()


def _rows_to_change_history_df(rows: list[list[str]]) -> pd.DataFrame:
    """
    Build a change-history DataFrame from a list of text rows.

    The first row is treated as the header; empty trailing rows are dropped and every row is
    padded/truncated to the header width. Used for the Office/HTML extraction paths where
    ``pdfplumber``'s page-spanning logic does not apply.
    """
    if len(rows) < 2:
        return pd.DataFrame()
    header = rows[0]
    width = len(header)
    data = [row for row in rows[1:] if any(cell for cell in row)]
    normalized = [(row + [""] * width)[:width] for row in data]
    if not normalized:
        return pd.DataFrame()
    return pd.DataFrame(normalized, columns=header)


def extract_change_history_from_xlsx(xlsx_path: Path) -> pd.DataFrame:
    """Extract the Änderungshistorie table from an ``.xlsx`` document, if it contains one."""
    try:
        workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("Failed to open XLSX %s: %s", xlsx_path.name, str(e))
        return pd.DataFrame()

    try:
        for worksheet in workbook.worksheets:
            rows = [[_cell_text(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
            for header_idx, row in enumerate(rows):
                if row and _is_change_history_header(row[0]):
                    df = _rows_to_change_history_df(rows[header_idx:])
                    if not df.empty:
                        logger.info("Extracted %d rows of change history data from %s", len(df), xlsx_path.name)
                        return df
    finally:
        workbook.close()

    logger.warning("No change history table found in %s", xlsx_path.name)
    return pd.DataFrame()


def extract_change_history_from_docx(docx_path: Path) -> pd.DataFrame:
    """Extract the Änderungshistorie table from a ``.docx`` document, reusing the docx pipeline."""
    try:
        document = docx.Document(str(docx_path))
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("Failed to open DOCX %s: %s", docx_path.name, str(e))
        return pd.DataFrame()

    change_history_table = get_change_history_table(document=document)
    if change_history_table is None:
        logger.warning("No change history table found in %s", docx_path.name)
        return pd.DataFrame()

    change_history_table.sanitize_table()
    return change_history_table.table


def extract_change_history_from_html(html_path: Path) -> pd.DataFrame:
    """Extract the Änderungshistorie table from a genuine HTML document, if it contains one."""
    soup = BeautifulSoup(html_path.read_text(encoding="utf-8", errors="replace"), "html.parser")
    for table in soup.find_all("table"):
        if not isinstance(table, Tag):
            continue
        rows: list[list[str]] = []
        for table_row in table.find_all("tr"):
            if not isinstance(table_row, Tag):
                continue
            rows.append([_cell_text(cell.get_text(" ", strip=True)) for cell in table_row.find_all(["td", "th"])])
        if rows and rows[0] and _is_change_history_header(rows[0][0]):
            df = _rows_to_change_history_df(rows)
            if not df.empty:
                logger.info("Extracted %d rows of change history data from %s", len(df), html_path.name)
                return df

    logger.warning("No change history table found in %s", html_path.name)
    return pd.DataFrame()


def extract_change_history_from_document(document_path: Path) -> pd.DataFrame:
    """
    Route a downloaded document to the extractor matching its (real) file type.

    Note: BNetzA serves many documents under a ``.html`` filename whose body is actually a PDF;
    those are stored with a ``.pdf`` extension at download time, so routing on the suffix here is
    correct.
    """
    suffix = document_path.suffix.lower()
    if suffix == ".pdf":
        return extract_change_history(document_path)
    if suffix == ".xlsx":
        return extract_change_history_from_xlsx(document_path)
    if suffix == ".docx":
        return extract_change_history_from_docx(document_path)
    if suffix in (".html", ".htm"):
        return extract_change_history_from_html(document_path)
    logger.warning("Unsupported document type for change history extraction: %s", document_path.name)
    return pd.DataFrame()


def _make_sheet_name(document_file: Path) -> str:
    """Create an Excel sheet name from a document filename, max 31 chars."""
    name = document_file.stem
    if "AHB" in name:
        name = f"AHB_{name.replace('_AHB', '')}"
    elif "MIG" in name:
        name = f"MIG_{name.replace('_MIG', '')}"
    if len(name) > 31:
        name = name[:28] + "..."
    return name


def _unique_sheet_name(name: str, used_names: set[str]) -> str:
    """Return a sheet name not already in ``used_names``, disambiguating with a numeric suffix."""
    if name not in used_names:
        return name
    for counter in range(2, 1000):
        suffix = f"_{counter}"
        candidate = name[: 31 - len(suffix)] + suffix
        if candidate not in used_names:
            return candidate
    return name[:31]


# Document extensions we know how to extract a change history from.
_SUPPORTED_EXTENSIONS = (".pdf", ".xlsx", ".docx", ".html", ".htm")


@dataclass
class ExtractionResult:
    """Result of extracting change histories from a set of downloaded documents."""

    sheets: list[tuple[str, pd.DataFrame]] = field(default_factory=list)
    no_change_history: list[str] = field(default_factory=list)  # processed, but no Änderungshistorie
    failed: list[str] = field(default_factory=list)  # raised while being processed


def _collect_sheets_data(document_files: list[Path]) -> ExtractionResult:
    """
    Extract change history data from documents.

    Documents that raise during processing are recorded in ``failed`` (a genuine error), kept
    distinct from ``no_change_history`` (documents that were read fine but simply contain no
    Änderungshistorie table).
    """
    result = ExtractionResult()
    used_names: set[str] = set()
    for document_file in sorted(document_files, key=lambda x: x.stem):
        try:
            if not document_file.is_file():
                logger.error("File %s is not a regular file", document_file.name)
                result.failed.append(document_file.name)
                continue
            logger.info("Processing %s (%d bytes)...", document_file.name, document_file.stat().st_size)

            df = extract_change_history_from_document(document_file)
            if not df.empty:
                sheet_name = _unique_sheet_name(_make_sheet_name(document_file), used_names)
                used_names.add(sheet_name)
                result.sheets.append((sheet_name, df))
                logger.info("Successfully extracted data from %s (%d rows)", document_file.name, len(df))
            else:
                result.no_change_history.append(document_file.name)
                logger.warning("No change history data found in %s", document_file.name)
        except Exception as e:  # pylint: disable=broad-exception-caught
            result.failed.append(document_file.name)
            logger.error("Failed to process %s: %s", document_file.name, str(e))
            continue

    logger.info(
        "Successfully extracted change histories from %d of %d documents",
        len(result.sheets),
        len(document_files),
    )
    return result


def _write_sheets_to_excel(sheets_data: list[tuple[str, pd.DataFrame]], output_file: Path) -> None:
    """Write collected sheets data to an Excel file with formatting."""
    output_file.parent.mkdir(parents=True, exist_ok=True)

    try:
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for sheet_name, df in sheets_data:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                worksheet = writer.sheets[sheet_name]

                column_widths = [10, 16, 33, 33, 30, 22]
                for idx, width in enumerate(column_widths):
                    worksheet.column_dimensions[chr(65 + idx)].width = width

                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True)

                logger.info("Successfully processed sheet %s", sheet_name)

        logger.info("Excel file created successfully at %s", output_file)
    except Exception as e:  # pylint: disable=broad-exception-caught
        logger.error("Failed to create Excel file: %s", str(e))
        raise


def create_change_history_excel(document_dir: Path, output_file: Path) -> ExtractionResult:
    """
    Create an Excel file containing change history tables from all documents in the directory.

    Args:
        document_dir: Directory containing the downloaded documents (PDF/Office/HTML)
        output_file: Path where the Excel file should be saved

    Returns:
        The :class:`ExtractionResult` describing which documents produced sheets, which contained
        no change history and which failed to process.
    """
    if not document_dir.exists():
        logger.error("Directory %s does not exist", document_dir)
        return ExtractionResult()

    document_files = [p for p in sorted(document_dir.iterdir()) if p.suffix.lower() in _SUPPORTED_EXTENSIONS]
    if not document_files:
        logger.warning("No documents found in directory %s", document_dir)
        return ExtractionResult()

    logger.info("Found %d documents to process", len(document_files))

    result = _collect_sheets_data(document_files)

    if not result.sheets:
        logger.warning("No change history data extracted from any document; no Excel file written")
        return result

    result.sheets.sort(key=lambda x: x[0])

    logger.info("Creating Excel file at %s with %d sheets", output_file, len(result.sheets))
    _write_sheets_to_excel(result.sheets, output_file)
    return result


@dataclass
class BNetzASummary:
    """Reconciliation of a BNetzA change-history run: links → downloads → extracted sheets."""

    links_found: int = 0
    downloaded: int = 0
    skipped: int = 0
    failed_downloads: list[str] = field(default_factory=list)
    kinds: dict[str, int] = field(default_factory=dict)  # by real type, downloaded files only
    sheets: list[str] = field(default_factory=list)
    no_change_history: list[str] = field(default_factory=list)
    failed_processing: list[str] = field(default_factory=list)
    output_file: Path | None = None


# Limit concurrent downloads to be polite to the BNetzA server and cap peak memory.
_MAX_CONCURRENT_DOWNLOADS = 8


async def download_documents(url: str, target_dir: Path | None = None) -> BNetzASummary:
    """
    Download all documents linked from the given BNetzA page and build the change-history Excel.

    Downloads PDFs, Office files and the ``.html``-named documents (which are usually PDFs),
    stores each under its real file type, extracts the Änderungshistorie from every document that
    has one, and writes one Excel sheet per document to ``<target_dir>/../change_history.xlsx``.

    Args:
        url: The BNetzA page URL to scrape for document links.
        target_dir: Directory to store downloaded documents. Defaults to a 'pdfs' directory next
            to this script.

    Returns:
        A :class:`BNetzASummary` reconciling links found, files downloaded and sheets extracted.
    """
    if target_dir is None:
        target_dir = Path(__file__).parent / "pdfs"

    target_dir.mkdir(parents=True, exist_ok=True)
    logger.info("Downloading documents to %s", target_dir)

    summary = BNetzASummary()

    document_links = await get_document_links(url)
    summary.links_found = len(document_links)
    if not document_links:
        logger.warning("No document links found on the page")
        return summary

    semaphore = asyncio.Semaphore(_MAX_CONCURRENT_DOWNLOADS)
    async with httpx.AsyncClient(timeout=httpx.Timeout(120.0), follow_redirects=True, headers=_HTTP_HEADERS) as client:
        logger.info("Starting download of %d documents...", len(document_links))
        results = await asyncio.gather(
            *(download_document(client, link_url, text, target_dir, semaphore) for link_url, text in document_links)
        )
        logger.info("Download process completed")

    for result in results:
        if result.status == "downloaded":
            summary.downloaded += 1
            # kinds describes freshly downloaded files, matching the "Downloaded" count shown.
            summary.kinds[result.kind] = summary.kinds.get(result.kind, 0) + 1
        elif result.status == "skipped":
            summary.skipped += 1
        elif result.status == "failed":
            summary.failed_downloads.append(result.stem)

    output_file = target_dir.parent / "change_history.xlsx"
    logger.info("Creating change history Excel file at %s", output_file)

    extraction = create_change_history_excel(target_dir, output_file)
    summary.sheets = [name for name, _ in extraction.sheets]
    summary.no_change_history = extraction.no_change_history
    summary.failed_processing = extraction.failed

    if extraction.sheets and output_file.exists():
        summary.output_file = output_file
        logger.info("Excel file created at %s (%d bytes)", output_file, output_file.stat().st_size)
    else:
        logger.error("No Excel file was created (no change history data extracted)")

    return summary


# Backwards-compatible alias; prefer :func:`download_documents`.
download_pdfs = download_documents
